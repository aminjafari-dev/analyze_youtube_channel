"""Excel handler for saving video data"""

import os
from datetime import datetime
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from src.data.models import VideoData
from src.exceptions.custom_exceptions import ExcelError
from src.utils.validators import sanitize_filename
from src.utils.config import EXCEL_COLUMNS


class ExcelHandler:
    """Handler for Excel file operations"""
    
    def __init__(self):
        """Initialize Excel handler"""
        self.workbook = None
        self.worksheet = None
    
    def create_file_name(self, channel_name: str) -> str:
        """Create filename with channel name, date, and time"""
        # Sanitize channel name
        safe_channel_name = sanitize_filename(channel_name)
        
        # Get current date and time
        now = datetime.now()
        date_str = now.strftime("%Y-%m-%d")
        time_str = now.strftime("%H-%M-%S")
        
        # Create filename
        filename = f"{safe_channel_name}_{date_str}_{time_str}.xlsx"
        
        return filename
    
    def create_workbook(self):
        """Create a new workbook"""
        try:
            self.workbook = Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.title = "YouTube Videos"
        except Exception as e:
            raise ExcelError(f"Failed to create workbook: {str(e)}")
    
    def setup_headers(self):
        """Setup column headers with styling"""
        try:
            # Header style
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Write headers
            for col_idx, header in enumerate(EXCEL_COLUMNS, 1):
                cell = self.worksheet.cell(row=1, column=col_idx)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Auto-adjust column widths
            self._adjust_column_widths()
            
        except Exception as e:
            raise ExcelError(f"Failed to setup headers: {str(e)}")
    
    def _adjust_column_widths(self):
        """Adjust column widths for better readability"""
        # Column width mappings (in characters)
        column_widths = {
            'Channel Name': 20,
            'Video Title': 40,
            'Video URL': 50,
            'Upload Date': 15,
            'Views': 15,
            'Likes': 15,
            'Comments': 15,
            'Description': 60,
            'Transcript': 80,
        }
        
        for col_idx, header in enumerate(EXCEL_COLUMNS, 1):
            width = column_widths.get(header, 20)
            self.worksheet.column_dimensions[get_column_letter(col_idx)].width = width
    
    def add_video_data(self, video: VideoData):
        """Add a single video's data to the worksheet"""
        try:
            # Get next row
            next_row = self.worksheet.max_row + 1
            
            # Get video data as dictionary
            data = video.to_dict()
            
            # Write data
            for col_idx, header in enumerate(EXCEL_COLUMNS, 1):
                cell = self.worksheet.cell(row=next_row, column=col_idx)
                value = data.get(header, '')
                
                # Format URL cells
                if header == 'Video URL' and value:
                    cell.hyperlink = value
                    cell.value = value
                    cell.font = Font(color="0000FF", underline="single")
                else:
                    cell.value = value
                
                # Set alignment
                if header in ['Views', 'Likes', 'Comments', 'Upload Date']:
                    cell.alignment = Alignment(horizontal="center", vertical="top")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            
        except Exception as e:
            raise ExcelError(f"Failed to add video data: {str(e)}")
    
    def add_videos(self, videos: List[VideoData]):
        """Add multiple videos to the worksheet"""
        try:
            for video in videos:
                self.add_video_data(video)
        except Exception as e:
            raise ExcelError(f"Failed to add videos: {str(e)}")
    
    def save(self, filepath: str):
        """Save workbook to file"""
        try:
            # Create directory if it doesn't exist
            directory = os.path.dirname(filepath)
            if directory and not os.path.exists(directory):
                os.makedirs(directory, exist_ok=True)
            
            # Save file
            self.workbook.save(filepath)
        except Exception as e:
            raise ExcelError(f"Failed to save file: {str(e)}")
    
    def export_videos(self, videos: List[VideoData], channel_name: str, output_dir: str = ".") -> str:
        """Export videos to Excel file"""
        try:
            # Create workbook
            self.create_workbook()
            
            # Setup headers
            self.setup_headers()
            
            # Add videos
            self.add_videos(videos)
            
            # Create filename
            filename = self.create_file_name(channel_name)
            filepath = os.path.join(output_dir, filename)
            
            # Save file
            self.save(filepath)
            
            return filepath
            
        except Exception as e:
            raise ExcelError(f"Failed to export videos: {str(e)}")

